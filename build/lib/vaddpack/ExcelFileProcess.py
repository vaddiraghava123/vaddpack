from openpyxl import load_workbook


def excelFileProcessAndUpdate(excelFile):
    workbook = load_workbook(excelFile)
    worksheet = workbook.active
    return 'Worksheet details are :: rows and columns'+ str(worksheet.max_row) + ':' + str(worksheet.max_column)
